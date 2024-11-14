from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.db.models.functions import TruncMonth
from django.db.models import Sum
from django.contrib import messages
from django.shortcuts import redirect

import openpyxl
import logging

from ..models import Expense, Category, User


logger = logging.getLogger('Expenses')


@login_required
def getYearlyExpensesView(request, year):
    data = getYearlyExpenses(request.user, year)
    return JsonResponse(data)


@login_required
def getMonthlyExpensesView(request, year, month):
    data = getMonthlyExpenses(request.user, year, month)
    return JsonResponse(data)


def getYearlyExpenses(user, year):
    # Initialize data for each month (1 to 12)
    data = {month: 0 for month in range(1, 13)}

    # Filter expenses by user and year, group by month, and sum amounts
    expenses_by_month = (
        Expense.objects.filter(date__year=year)
        .annotate(month=TruncMonth('date'))
        .values('month')
        .annotate(total=Sum('amount'))
    )

    # Populate the data dictionary with the monthly sums
    for entry in expenses_by_month:
        month = entry['month'].month
        total = entry['total']
        data[month] = float(total)  # Convert Decimal to float for easier JSON serialization
        
    return data


def getMonthlyExpenses(user, year, month):
    categories = Category.objects.all()
    data = {category.name: 0 for category in categories}

    # Filter expenses by the year, month, and user
    expenses = Expense.objects.filter(date__year=year, date__month=month)
    
    # Sum the amounts for each category
    for expense in expenses:
        data[expense.category.name] += expense.amount  # Summing the expense amount
        
    return data


def importFromExcel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active

            errors = []  # To collect error messages

            # Assuming the columns are: Category, Amount, Currency, Date, Description, Is Paid, Split With
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                category, amount, currency, date, description, is_paid, split_with = row

                # Check if the user exists
                if split_with != "Both":
                    if split_with not in User.objects.all():
                        errors.append(f"User named {split_with} does not exist.")
                        continue
                # Check if category exists
                if category not in Category.objects.all():
                    errors.append(f"Category named: {category} does not exist.")
                    continue                    

                Expense.objects.create(
                    user=request.user,  # Assuming imported expenses are for the logged-in user
                    category=category,
                    amount=amount,
                    currency=currency,
                    date=date,
                    description=description,
                    isPaid=(str(is_paid).lower() == 'yes'),
                    splitWith=split_with
                )

            # Display messages
            if errors:
                for error in errors:
                    messages.error(request, error)
            else:
                messages.success(request, "Expenses imported successfully!")

            return JsonResponse({'message': 'Expenses imported successfully.'})
        
        except Exception as e:
            messages.error(request, f"Error importing expenses: {str(e)}")

        return redirect('expenses')

    else:
        messages.error(request, "Invalid request.")
        return redirect('expenses')


def exportToExcel(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Expenses'

    # Define the headers
    headers = ['Category', 'Amount', 'Currency', 'Date', 'Description', 'Is Paid', 'Split With']
    worksheet.append(headers)

    # Populate the worksheet with data
    expenses = Expense.objects.all()
    for expense in expenses:
        worksheet.append([
            expense.category.name,
            expense.amount,
            expense.currency,
            expense.date,
            expense.description,
            'Yes' if expense.isPaid else 'No',
            expense.get_split_with_display()
        ])

    # Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=expenses_export.xlsx'
    workbook.save(response)
    return response